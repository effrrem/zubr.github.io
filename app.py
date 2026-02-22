import os
import smtplib
from flask import Flask, render_template, request, redirect, url_for, session
import sqlite3
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

app = Flask(__name__)
app.secret_key = 'supersecretkey'

with app.app_context():
    import sqlite3
    conn = sqlite3.connect("database.db")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


# Проверка, авторизован ли админ
def is_admin_logged_in():
    return session.get('admin_logged_in') is True


@app.route("/", methods=["GET", "POST"])
def index():
    search_query = request.form.get("search_query", "").strip() if request.method == "POST" else ""
    search_type = request.form.get("search_type", "name") if request.method == "POST" else "name"

    parts = []
    query = """
        SELECT id, name, catalog_number, retail_price, stock, description 
        FROM parts 
        WHERE 1=1
    """
    params = []

    with sqlite3.connect("database.db") as conn:
        # Регистронезависимый поиск (работает с кириллицей)
        conn.create_function("LOWER", 1, lambda s: str(s).lower() if s else "")

        if search_query:
            if search_type == "name":
                query += " AND LOWER(name) LIKE ?"
                params.append(f"%{search_query.lower()}%")
            elif search_type == "catalog":
                query += " AND LOWER(catalog_number) LIKE ?"
                params.append(f"%{search_query.lower()}%")
            elif search_type == "price":
                try:
                    price = float(search_query)
                    query += " AND retail_price = ?"
                    params.append(price)
                except ValueError:
                    pass
            elif search_type == "stock":
                if search_query.isdigit():
                    query += " AND stock >= ?"
                    params.append(int(search_query))

        cursor = conn.execute(query, params)
        parts = cursor.fetchall()

    return render_template(
        "index.html",
        parts=parts,
        search_query=search_query,
        search_type=search_type
    )


@app.route("/about")
def about():
    return render_template("about.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        if username == "ЕфремовАВ900" and password == "27051984":
            session['admin_logged_in'] = True
            return redirect(url_for("admin"))
        else:
            return render_template("login.html", error="Неверный логин или пароль")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("index"))


@app.route("/admin")
def admin():
    if not is_admin_logged_in():
        return redirect(url_for("login"))
    with sqlite3.connect("database.db") as conn:
        cursor = conn.execute("""
            SELECT id, name, catalog_number, retail_price, stock 
            FROM parts
        """)
        parts = cursor.fetchall()
    return render_template("admin.html", parts=parts)


@app.route("/add", methods=["GET", "POST"])
def add_part():
    if not is_admin_logged_in():
        return redirect(url_for("login"))

    if request.method == "POST":
        name = request.form["name"]
        catalog_number = request.form.get("catalog_number", "").strip()
        try:
            retail_price = float(request.form["retail_price"])
            if retail_price < 0:
                return "Цена не может быть отрицательной", 400
        except (ValueError, TypeError):
            return "Неверная цена", 400

        stock = request.form.get("stock", "0").strip()
        try:
            stock = int(stock) if stock else 0
            if stock < 0:
                return "Количество на складе не может быть отрицательным", 400
        except (ValueError, TypeError):
            return "Неверное значение склада", 400

        description = request.form.get("description", "").strip()

        try:
            with sqlite3.connect("database.db") as conn:
                conn.execute("""
                    INSERT INTO parts (name, catalog_number, retail_price, stock, description)
                    VALUES (?, ?, ?, ?, ?)
                """, (name, catalog_number, retail_price, stock, description))
            return redirect(url_for("admin"))
        except sqlite3.IntegrityError:
            return "Ошибка: № по каталогу уже существует!", 400

    return render_template("add_part.html")


@app.route("/delete/<int:part_id>")
def delete_part(part_id):
    if not is_admin_logged_in():
        return redirect(url_for("login"))
    with sqlite3.connect("database.db") as conn:
        conn.execute("DELETE FROM parts WHERE id = ?", (part_id,))
    return redirect(url_for("admin"))


from openpyxl import load_workbook
from werkzeug.utils import secure_filename


@app.route("/upload_excel", methods=["POST"])
def upload_excel():
    if not is_admin_logged_in():
        return redirect(url_for("login"))
    if 'excel_file' not in request.files:
        return "Файл не выбран", 400
    file = request.files['excel_file']
    if file.filename == '':
        return "Файл не выбран", 400
    if file and file.filename.endswith(('.xlsx', '.xls')):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        try:
            workbook = load_workbook(filepath)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            expected = ['Наименование', '№ по каталогу', 'Розничная', 'Склад']
            if headers != expected:
                return f"Неверные заголовки. Нужны: {expected}", 400

            with sqlite3.connect("database.db") as conn:
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    name = str(row[0]).strip() if row[0] else ""
                    catalog_number = str(row[1]).strip() if row[1] else ""

                    # Обработка цены
                    price_val = row[2]
                    if not price_val:
                        retail_price = 0.0
                    else:
                        try:
                            price_str = str(price_val).strip()
                            retail_price = float(price_str)
                            if retail_price < 0:
                                retail_price = 0.0
                        except (ValueError, TypeError):
                            print(f"⚠️ Строка {row_num}: невозможно преобразовать '{price_val}' в цену → 0.0")
                            retail_price = 0.0

                    # Обработка склада
                    stock_val = row[3]
                    if not stock_val:
                        stock = 0
                    else:
                        try:
                            stock_str = str(stock_val).strip()
                            stock = int(float(stock_str))
                            if stock < 0:
                                stock = 0
                        except (ValueError, TypeError):
                            print(f"⚠️ Строка {row_num}: некорректное значение склада '{stock_val}' → 0")
                            stock = 0

                    description = ""

                    if not name:
                        print(f"⚠️ Пропущено: нет названия (строка {row_num})")
                        continue

                    try:
                        conn.execute("""
                            INSERT INTO parts (name, catalog_number, retail_price, stock, description)
                            VALUES (?, ?, ?, ?, ?)
                        """, (name, catalog_number, retail_price, stock, description))
                    except sqlite3.IntegrityError:
                        print(f"⚠️ Дубликат: {name} (строка {row_num})")
                        pass

            if os.path.exists(filepath):
                os.remove(filepath)
            return redirect(url_for("admin"))
        except Exception as e:
            if os.path.exists(filepath):
                os.remove(filepath)
            return f"Ошибка при обработке файла: {str(e)}", 500
    else:
        return "Неверный формат файла", 400


import hashlib

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        confirm_password = request.form["confirm_password"]
        if password != confirm_password:
            return "Пароли не совпадают", 400
        hashed = hash_password(password)
        try:
            with sqlite3.connect("database.db") as conn:
                conn.execute(
                    "INSERT INTO users (username, password) VALUES (?, ?)",
                    (username, hashed)
                )
            return redirect(url_for("login_user"))
        except sqlite3.IntegrityError:
            return "Пользователь с таким именем уже существует", 400
    return render_template("register.html")


@app.route("/login_user", methods=["GET", "POST"])
def login_user():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        hashed = hash_password(password)
        with sqlite3.connect("database.db") as conn:
            cursor = conn.execute(
                "SELECT id FROM users WHERE username = ? AND password = ?",
                (username, hashed)
            )
            user = cursor.fetchone()
        if user:
            session['user_id'] = user[0]
            session['username'] = username
            return redirect(url_for("index"))
        else:
            return "Неверный логин или пароль", 401
    return render_template("login_user.html")


@app.route("/logout_user")
def logout_user():
    session.clear()
    return redirect(url_for("index"))


def send_service_notification(name, phone, email, car_model, service_type, date, comment):
    SMTP_SERVER = "smtp.mail.ru"
    SMTP_PORT = 465
    SMTP_USERNAME = "matvey.efremov@internet.ru"
    SMTP_PASSWORD = "j3WeS7cmJj2XXBzfJhT8"

    msg = MIMEMultipart()
    msg["From"] = SMTP_USERNAME
    msg["To"] = "efremov@tehcentr.ru"
    msg["Subject"] = f"Новая запись на сервис: {name}"

    body = f"""
Новая запись на сервисное обслуживание!

Имя: {name}
Телефон: {phone}
Email: {email}
Модель автомобиля: {car_model}
Вид услуги: {service_type}
Желаемая дата: {date}
Комментарий: {comment}
"""
    msg.attach(MIMEText(body, "plain"))

    try:
        server = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT, timeout=10)
        server.login(SMTP_USERNAME, SMTP_PASSWORD)
        server.sendmail(SMTP_USERNAME, "efremov@tehcentr.ru", msg.as_string())
        server.quit()
        return True
    except Exception as e:
        print(f"Email error: {e}")
        return False


@app.route("/service", methods=["GET", "POST"])
def service():
    success = False
    if request.method == "POST":
        name = request.form.get("name")
        phone = request.form.get("phone")
        email = request.form.get("email")
        car_model = request.form.get("car_model")
        service_type = request.form.get("service_type")
        date = request.form.get("date")
        comment = request.form.get("comment")

        success = send_service_notification(name, phone, email, car_model, service_type, date, comment)

    return render_template("service.html", success=success)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)